# skills/data_engine/artifacts.py
# ──────────────────────────────────────────────────────────────────────
# Report artifacts for the deep lane: Markdown, Excel, and rerunnable
# Jupyter notebook — all generated from COMPUTED results (profile,
# findings, insights), never fresh LLM output. Builders never fail the
# answer: voice delivers even if a builder errors (caller guards).
# ──────────────────────────────────────────────────────────────────────

# 1. stdlib
import datetime
import json
import logging
from pathlib import Path

logger = logging.getLogger('dna.data_engine.artifacts')


def _stamp() -> str:
    return datetime.datetime.now().strftime("%Y%m%d_%H%M%S")


def build_markdown(profile: dict, findings: list, insights: dict,
                   history: list, out_dir, meta: dict | None = None) -> str:
    """Write findings report markdown. Returns path or ''."""
    try:
        meta = meta or {}
        name = Path(profile.get('file_path', 'dataset')).name
        lines = [
            f"# Data Report — {name}",
            "",
            f"_Generated {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')} | "
            f"{profile.get('row_count', 0):,} rows × {profile.get('column_count', 0)} cols | "
            f"quality {profile.get('quality_score', 0.0):.1f}% | "
            f"mode: {'local-only' if meta.get('local_only') else 'cloud'} | "
            f"question: {meta.get('question', '')}_",
            "",
            "## Executive summary",
            "",
            insights.get('executive_summary', 'n/a'),
            "",
            "## KPIs",
            "",
        ]
        for kpi in insights.get('kpis', []) or []:
            lines.append(f"- **{kpi.get('label', '')}**: {kpi.get('value', '')} — {kpi.get('detail', '')}")
        lines += ["", "## Key drivers", ""]
        for d in insights.get('drivers', []) or []:
            lines.append(f"- **{d.get('title', '')}** [{d.get('severity', '')}]: {d.get('business_insight', '')}")
        lines += ["", "## Outliers & anomalies", ""]
        for o in insights.get('outliers_and_anomalies', []) or []:
            lines.append(f"- **{o.get('column', '')}** [{o.get('severity', '')}]: "
                         f"{o.get('outlier_count', '')} outliers, {o.get('business_insight', '')}")
        lines += ["", "## Recommendations", ""]
        for r in insights.get('recommendations', []) or []:
            lines.append(f"- **{r.get('title', '')}**: {r.get('action', '')} _{r.get('rationale', '')}_")
        if findings:
            lines += ["", "## Detector findings", ""]
            for f in findings[:20]:
                lines.append(f"- [{f.get('severity', '')}] {f.get('detail', '')}")
        if history:
            lines += ["", "## Prior questions on this dataset", ""]
            for h in history[-10:]:
                lines.append(f"- {h.get('question', '')}")
        out = Path(out_dir) / f"report_{_stamp()}.md"
        out.write_text("\n".join(lines) + "\n", encoding="utf-8")
        return str(out)
    except Exception as e:
        logger.warning('Markdown build failed: %s', e)
        return ""


def build_excel(profile: dict, findings: list, insights: dict,
                chart_paths: list, history: list, out_dir) -> str:
    """Write .xlsx findings workbook with embedded chart PNGs. Returns path or ''."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.drawing.image import Image as XLImage
    except Exception as e:
        logger.warning('openpyxl unavailable, skipping Excel: %s', e)
        return ""
    try:
        wb = Workbook()
        bold = Font(bold=True)

        def _sheet(title, rows):
            ws = wb.create_sheet(title[:31]) if title != "Sheet" else wb.active
            if title == "Sheet":
                ws.title = title[:31]
            for r in rows:
                ws.append([str(c) for c in r])
            for cell in ws[1]:
                cell.font = bold
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = min(60, max(14, max(
                    (len(str(c.value or "")) for c in col), default=14)))
            return ws

        name = Path(profile.get('file_path', 'dataset')).name
        _sheet("Sheet", [["Metric", "Value"],
                         ["Dataset", name],
                         ["Rows", profile.get('row_count', 0)],
                         ["Columns", profile.get('column_count', 0)],
                         ["Quality %", round(float(profile.get('quality_score', 0.0)), 1)],
                         ["Summary", (insights.get('executive_summary', '') or '')[:500]]])
        _sheet("Findings", [["Severity", "Type", "Column", "Detail"]] + [
            [f.get('severity', ''), f.get('type', ''), f.get('column', ''), f.get('detail', '')]
            for f in (findings or [])[:100]])
        _sheet("Recommendations", [["Title", "Action", "Rationale"]] + [
            [r.get('title', ''), r.get('action', ''), r.get('rationale', '')]
            for r in (insights.get('recommendations', []) or [])])
        _sheet("Schema", [["Column", "Type", "Null %"]] + [
            [c.get('name', ''), c.get('type', ''),
             round(float((profile.get('null_summary', {}) or {}).get(c.get('name', ''), 0) or 0), 1)]
            for c in (profile.get('schema', []) or [])])
        charts_ws = _sheet("Charts", [["Chart file"]])
        row = 3
        for png in (chart_paths or [])[:6]:
            try:
                charts_ws.append([Path(png).name])
                img = XLImage(str(png))
                img.width, img.height = 640, 360
                charts_ws.add_image(img, f"B{row}")
                row += 20
            except Exception as e:
                logger.debug('Chart embed skipped for %s: %s', png, e)
        out = Path(out_dir) / f"report_{_stamp()}.xlsx"
        wb.save(str(out))
        return str(out)
    except Exception as e:
        logger.warning('Excel build failed: %s', e)
        return ""


def _nb_cell(cell_type: str, source: list[str], outputs=None, lang="python") -> dict:
    cell = {"cell_type": cell_type, "metadata": {}, "source": source}
    if cell_type == "code":
        cell["execution_count"] = None
        cell["outputs"] = outputs or []
    return cell


def _nb_out(text: str) -> dict:
    return {"output_type": "stream", "name": "stdout", "text": [text[:2000]]}


def build_notebook(profile: dict, question: str, query_log: list,
                   history: list, out_dir, filename: str = "dataset.csv") -> str:
    """Write rerunnable .ipynb reproducing the analysis (hand-rolled nbformat 4).

    Code cells contain the ACTUAL executed SQL plus reproduce-from-scratch
    pandas cells, so colleagues rerun without DNA. No nbformat dependency.
    """
    try:
        name = Path(profile.get('file_path', 'dataset')).name
        schema = profile.get('schema') or []
        col_list = ", ".join(c.get('name', '') for c in schema[:12])
        cells = [
            _nb_cell("markdown", [
                f"# Reproducible analysis — {name}\n",
                f"Question: {question}\n",
                f"Generated by DNA deep lane. Point FILE at your copy and run all.\n",
            ]),
            _nb_cell("code", [
                "# pip install duckdb pandas matplotlib openpyxl\n",
                "import duckdb\nimport pandas as pd\nimport matplotlib.pyplot as plt\n",
                f'FILE = "{filename}"  # <-- point at your copy of the dataset\n',
                "con = duckdb.connect()\n",
                "print('ready')\n",
            ], outputs=[_nb_out("ready\n")]),
            _nb_cell("code", [
                f"df = con.execute(\"SELECT * FROM read_csv_auto('\" + FILE + \"')\").fetchdf() "
                "# swap reader per format: read_parquet / pd.read_excel\n",
                "print(df.shape)\nprint(df.dtypes)\n",
                "df.head()\n",
            ]),
            _nb_cell("code", [
                "print('nulls per column:')\nprint(df.isna().sum()[df.isna().sum() > 0])\n",
                "print(df.describe(include='all').T)\n",
            ]),
        ]
        for desc, sql in (query_log or [])[-6:]:
            cells.append(_nb_cell("code", [
                f"# {desc}\n",
                f"result = con.execute(\"\"\"{sql}\"\"\").fetchdf()\n",
                "print(result.head(10).to_string())\n",
            ]))
        num_cols = [c.get('name', '') for c in schema
                    if 'int' in str(c.get('type', '')).lower()
                    or 'double' in str(c.get('type', '')).lower()
                    or 'float' in str(c.get('type', '')).lower()][:3]
        if num_cols:
            cells.append(_nb_cell("code", [
                f"df[{num_cols!r}].hist(bins=30, figsize=(10, 4))\nplt.tight_layout()\nplt.show()\n",
            ]))
        cells.append(_nb_cell("markdown", [
            "## Columns\n", f"{col_list}\n",
            "## Rerun\n", "Run all cells top to bottom against your copy of the file.\n",
        ]))
        nb = {"nbformat": 4, "nbformat_minor": 5,
              "metadata": {"language_info": {"name": "python"}},
              "cells": cells}
        out = Path(out_dir) / f"analysis_{_stamp()}.ipynb"
        out.write_text(json.dumps(nb, indent=1), encoding="utf-8")
        return str(out)
    except Exception as e:
        logger.warning('Notebook build failed: %s', e)
        return ""
